import bs4
import requests
import openpyxl
import math

lista_otros_telefonos = []
lista_nombres_empresas = []
lista_webs = []
lista_link_subpaginas = []

def crear_sopa(link):
    pedido = requests.get(link)

    sopa = bs4.BeautifulSoup(pedido.text, 'lxml')
    return sopa


def encontrar_datos(sopa):
    global lista_nombres_empresas
    global lista_webs
    global lista_link_subpaginas
    numero_link_subpaginas_extraidos_esta_pagina = 0
    sublista_nombres_empresas = []
    c = 1

    # Buscamos y agregamos nombres de empresa a lista
    elementos_span_propiedad_nombre = sopa.find_all("span", {"itemprop": "name"})
    for elemento in elementos_span_propiedad_nombre:
        sublista_nombres_empresas.append(elemento.text.strip())
        c += 1
    sublista_nombres_empresas = sublista_nombres_empresas[4:]


    for elemento in sublista_nombres_empresas:
        lista_nombres_empresas.append(elemento)



    # Buscamos y guardamos links de webs y links de subpáginas
    elementos_td_webs = sopa.find_all("td")

    contador = 1
    for elemento in elementos_td_webs:

        # Guardamos links de webs
        if contador % 4 == 0:
            lista_webs.append(elemento.text.strip())

        # Guardamos links de subpáginas
        etiqueta_a = elemento.find("a")
        if etiqueta_a is not None and numero_link_subpaginas_extraidos_esta_pagina < 20:
            enlace_subpagina = etiqueta_a['href']
            lista_link_subpaginas.append(enlace_subpagina)
            numero_link_subpaginas_extraidos_esta_pagina += 1

        contador += 1



def encontrar_datos_subpaginas(link_subpagina):
    global lista_otros_telefonos

    # Buscamos teléfonos
    sopa_subpagina = crear_sopa(link_subpagina)
    telefono = sopa_subpagina.find("span", {"class": "tel"})

    # Buscamos Otros Teléfonos
    hay_otros_telefonos = False
    etiqueta_table = sopa_subpagina.find("table", {"class": "vcard datos_ppales", "id": "table-datos-pples"})
    if etiqueta_table:
        etiquetas_tr = etiqueta_table.find_all("tr")



        for etiqueta_tr in etiquetas_tr:
            etiqueta_th = etiqueta_tr.find("th")
            if etiqueta_th.text == "Otros teléfonos:":
                hay_otros_telefonos = True
                otro_telefono = etiqueta_tr.find("span").text
                if len(otro_telefono) > 9:
                    otro_telefono = f"{otro_telefono[:9]} - {otro_telefono[9:]}"

    if hay_otros_telefonos:
        lista_otros_telefonos.append(otro_telefono)
    else:
        lista_otros_telefonos.append("")




    if telefono is not None:
        return telefono.text
    else:
        return ""




# Pedimos Datos al Usuario
link = input("Introduce el link ")
resultados = input("Cuantos resultados tiene la busqueda? ")
nombre_archivo = input("Como quieres llamar al archivo?")

# Calculamos cuantas páginas hay que recorrer y hacemos un bucle sacando los datos principales de cada una
resultados = math.ceil(int(resultados) / 20)
for n in range(1, resultados + 1):
    nuevo_link = f"{link[:-5]}/Empresas-{n}.html"

    # Creamos La sopa Principal
    sopa = crear_sopa(nuevo_link)

    # Sacamos Datos de Pagina Principal (nombre empresa, web, subpagina)
    encontrar_datos(sopa)

print(len(lista_nombres_empresas))
print(len(lista_webs))
print(len(lista_link_subpaginas))

# Sacamos los teléfonos de cada subpágina
lista_telefonos = []
for link_subpagina in lista_link_subpaginas:
    telefono = encontrar_datos_subpaginas(link_subpagina)
    lista_telefonos.append(telefono)

print(lista_telefonos)
print(len(lista_telefonos))
print(lista_otros_telefonos)
print(len(lista_otros_telefonos))

# Creamos Documento y Hoja
workbook = openpyxl.Workbook()
hoja = workbook.active

# Agregamos Nombres Columnas en 1ª fila
nombres_colum = ["Empresa", "Web", "Teléfono", "Otros teléfonos"]
for num_colum, nombre_colum in enumerate(nombres_colum, 1):
    cell = hoja.cell(row=1, column=num_colum)
    cell.value = nombre_colum

# Metemos todos los datos en una lista
lista_datos_final = [lista_nombres_empresas, lista_webs, lista_telefonos, lista_otros_telefonos]

# Pasamos datos a hoja Excel
for num_colum, lista_datos in enumerate(lista_datos_final, 1):
    for num_fila, dato in enumerate(lista_datos, 2):
        cell = hoja.cell(row=num_fila, column=num_colum)
        cell.value = dato

# Guardamos Documento
workbook.save(f"{nombre_archivo}.xlsx")

# Cerramos Documento
workbook.close()

